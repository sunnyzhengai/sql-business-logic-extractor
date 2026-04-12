#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SQL Business Logic Extractor -- Main Entry Point

Batch processes SQL files and generates:
1. Detailed JSON/text files per report (L3 lineage + L4 English definitions)
2. Summary Excel spreadsheet grouped by business logic for steward curation

Orchestrates: L3 (resolve) → L4 (translate) → L5 (compare) → Excel export

Usage:
    python3 governance_extract.py ./sql_reports/ --output governance_summary.xlsx
"""

import glob
import json
import os
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

# Import existing modules
from resolve import resolve_query, resolved_to_dict
from compare_lineage import LineageComparator, ResolvedDefinition


@dataclass
class BusinessLogicGroup:
    """A group of definitions for the same business logic term."""
    term_name: str
    status: str  # CONFLICT, SIMILAR, CONSISTENT, UNIQUE
    variation_count: int
    definitions: list[dict] = field(default_factory=list)


@dataclass
class ReportSummary:
    """Summary of a single report."""
    report_name: str
    report_description: str  # From L6 summarize_query
    primary_purpose: str
    column_count: int
    source_tables: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Batch Processing
# ---------------------------------------------------------------------------

def process_sql_file(sql_path: str, schema: dict, client, dialect: str = None) -> dict:
    """Process a single SQL file through L5 and L6.

    Returns:
        {
            'report_name': str,
            'l5_data': dict,  # Resolved lineage
            'l6_data': dict,  # English definitions + summary
        }
    """
    report_name = os.path.splitext(os.path.basename(sql_path))[0]

    # Read SQL
    with open(sql_path, 'r') as f:
        sql = f.read()

    # L5: Resolve lineage
    print(f"  [{report_name}] Resolving lineage...")
    resolved = resolve_query(sql, dialect=dialect)
    l5_data = resolved_to_dict(resolved)

    # L6: Translate to English (if client provided)
    l6_data = None
    if client and schema:
        from llm_translate import translate_column, summarize_query, build_column_context

        print(f"  [{report_name}] Translating to English...")
        columns = l5_data.get('columns', [])
        column_results = []

        for col in columns:
            result = translate_column(col, schema, client)
            column_results.append(result)

        # Generate query summary
        summary = summarize_query(column_results, l5_data, client)

        l6_data = {
            'summary': summary,
            'columns': column_results
        }

    return {
        'report_name': report_name,
        'sql_path': sql_path,
        'l5_data': l5_data,
        'l6_data': l6_data,
    }


def batch_process(sql_folder: str, schema_path: str = None, api_key: str = None,
                  dialect: str = None, output_dir: str = None) -> list[dict]:
    """Process all SQL files in a folder.

    Args:
        sql_folder: Path to folder containing SQL files
        schema_path: Path to clarity_schema.yaml (optional, for L6)
        api_key: OpenAI API key (optional, for L6)
        dialect: SQL dialect
        output_dir: Directory to save individual L5/L6 outputs

    Returns:
        List of processed report dicts
    """
    # Find SQL files
    sql_files = glob.glob(os.path.join(sql_folder, "*.sql"))
    if not sql_files:
        print(f"No SQL files found in {sql_folder}")
        return []

    print(f"Found {len(sql_files)} SQL files")

    # Load schema if provided
    schema = None
    client = None
    if schema_path and api_key:
        from llm_translate import load_schema
        from openai import OpenAI

        print(f"Loading schema: {schema_path}")
        schema = load_schema(schema_path)
        client = OpenAI(api_key=api_key)
    elif schema_path or api_key:
        print("Warning: Both --schema and --api-key required for L6 translation")

    # Create output directory
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        os.makedirs(os.path.join(output_dir, "details"), exist_ok=True)

    # Process each file
    results = []
    for i, sql_path in enumerate(sorted(sql_files), 1):
        report_name = os.path.splitext(os.path.basename(sql_path))[0]
        print(f"\n[{i}/{len(sql_files)}] Processing: {report_name}")

        try:
            result = process_sql_file(sql_path, schema, client, dialect)
            results.append(result)

            # Save individual outputs
            if output_dir:
                detail_path = os.path.join(output_dir, "details", report_name)

                # Save L5
                with open(f"{detail_path}_L5.json", 'w') as f:
                    json.dump(result['l5_data'], f, indent=2)

                # Save L6 if available
                if result['l6_data']:
                    with open(f"{detail_path}_L6.json", 'w') as f:
                        json.dump(result['l6_data'], f, indent=2)

        except Exception as e:
            print(f"  Error processing {report_name}: {e}")
            continue

    return results


# ---------------------------------------------------------------------------
# Comparison & Grouping
# ---------------------------------------------------------------------------

def compare_all_definitions(results: list[dict]) -> dict:
    """Compare all definitions across reports to find conflicts/duplicates.

    Returns comparison report dict.
    """
    comparator = LineageComparator()

    for result in results:
        report_name = result['report_name']
        l5_data = result['l5_data']

        # Add each column as a definition
        for col in l5_data.get('columns', []):
            defn = ResolvedDefinition(
                id=f"{report_name}:{col['name']}",
                query_label=report_name,
                column_name=col['name'],
                column_type=col.get('type', 'unknown'),
                resolved_expression=col.get('resolved_expression', ''),
                base_tables=col.get('base_tables', []),
                base_columns=col.get('base_columns', []),
                filters=col.get('filters', []),
            )
            comparator.definitions.append(defn)

        comparator.query_labels.add(report_name)

    # Run comparison
    report = comparator.compare(skip_trivial=True)

    return report


def group_by_business_logic(results: list[dict], comparison_report) -> list[BusinessLogicGroup]:
    """Group all definitions by business logic term.

    Each group contains:
    - All variations of the same term across reports
    - Status: CONFLICT, SIMILAR, CONSISTENT, UNIQUE
    """
    # Build index: term_name -> list of (report_name, definition)
    term_index = defaultdict(list)

    for result in results:
        report_name = result['report_name']
        l5_data = result['l5_data']
        l6_data = result.get('l6_data', {})

        # Build L6 lookup by column name
        l6_columns = {}
        if l6_data:
            for col in l6_data.get('columns', []):
                l6_columns[col['column_name']] = col

        # Index each column
        for col in l5_data.get('columns', []):
            col_name = col['name']
            col_type = col.get('type', 'unknown')

            # Skip trivial columns
            if col_type in ('passthrough', 'star', 'literal'):
                continue

            # Get L6 translation if available
            l6_col = l6_columns.get(col_name, {})

            term_index[col_name.lower()].append({
                'report_name': report_name,
                'column_name': col_name,
                'column_type': col_type,
                'resolved_expression': col.get('resolved_expression', ''),
                'base_tables': col.get('base_tables', []),
                'base_columns': col.get('base_columns', []),
                'filters': col.get('filters', []),
                'business_definition': l6_col.get('english_definition', ''),
                'technical_definition': col.get('resolved_expression', ''),
                'business_domain': l6_col.get('business_domain', ''),
            })

    # Build conflict set for quick lookup
    conflict_names = set()
    for conflict in comparison_report.conflicts:
        conflict_names.add(conflict.column_name.lower())

    # Build exact duplicate signatures
    exact_sigs = defaultdict(set)
    for group in comparison_report.exact_duplicates:
        for d in group.definitions:
            exact_sigs[d['column_name'].lower()].add(group.signature)

    # Build structural match signatures
    structural_sigs = defaultdict(set)
    for group in comparison_report.structural_matches:
        for d in group.definitions:
            structural_sigs[d['column_name'].lower()].add(group.signature)

    # Create groups
    groups = []
    for term_name, definitions in sorted(term_index.items()):
        # Determine status
        if term_name in conflict_names:
            status = "CONFLICT"
        elif len(definitions) > 1:
            # Check if all definitions are exact duplicates
            if term_name in exact_sigs:
                status = "CONSISTENT"
            elif term_name in structural_sigs:
                status = "SIMILAR"
            else:
                # Multiple definitions, need to check if they're the same
                expressions = set(d['resolved_expression'] for d in definitions)
                if len(expressions) == 1:
                    status = "CONSISTENT"
                else:
                    status = "SIMILAR"
        else:
            status = "UNIQUE"

        groups.append(BusinessLogicGroup(
            term_name=definitions[0]['column_name'],  # Use original case
            status=status,
            variation_count=len(definitions),
            definitions=definitions,
        ))

    # Sort: CONFLICT first, then SIMILAR, then CONSISTENT, then UNIQUE
    status_order = {"CONFLICT": 0, "SIMILAR": 1, "CONSISTENT": 2, "UNIQUE": 3}
    groups.sort(key=lambda g: (status_order.get(g.status, 99), -g.variation_count, g.term_name))

    return groups


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def export_to_excel(results: list[dict], groups: list[BusinessLogicGroup],
                    comparison_report, output_path: str):
    """Export governance summary to Excel spreadsheet.

    Sheet 1: Business Logic Summary (grouped by term)
    Sheet 2: Report Summary (one row per report)
    Sheet 3: All Definitions (flat list)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("Error: openpyxl required for Excel export. Install with: pip install openpyxl")
        # Fall back to CSV
        export_to_csv(results, groups, comparison_report, output_path)
        return

    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------------------
    # Sheet 1: Business Logic Summary
    # ---------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Business Logic"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    conflict_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    similar_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    consistent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Headers
    headers = [
        "Business Logic Term", "Status", "# Variations", "Report Name",
        "Business Definition", "Technical Definition", "Source Tables",
        "Business Domain", "Assigned To", "Review Status", "Notes"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Data rows
    row = 2
    for group in groups:
        first_row_of_group = row

        for i, defn in enumerate(group.definitions):
            ws1.cell(row=row, column=1, value=group.term_name if i == 0 else "")

            status_cell = ws1.cell(row=row, column=2, value=group.status if i == 0 else "")
            if group.status == "CONFLICT":
                status_cell.fill = conflict_fill
            elif group.status == "SIMILAR":
                status_cell.fill = similar_fill
            elif group.status == "CONSISTENT":
                status_cell.fill = consistent_fill

            ws1.cell(row=row, column=3, value=group.variation_count if i == 0 else "")
            ws1.cell(row=row, column=4, value=defn['report_name'])
            ws1.cell(row=row, column=5, value=defn.get('business_definition', ''))
            ws1.cell(row=row, column=6, value=defn.get('technical_definition', '')[:500])
            ws1.cell(row=row, column=7, value=", ".join(defn.get('base_tables', [])))
            ws1.cell(row=row, column=8, value=defn.get('business_domain', ''))
            # Columns 9-11 left empty for steward input

            row += 1

        # Add empty row between groups
        row += 1

    # Adjust column widths
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 50
    ws1.column_dimensions['F'].width = 60
    ws1.column_dimensions['G'].width = 30
    ws1.column_dimensions['H'].width = 20
    ws1.column_dimensions['I'].width = 15
    ws1.column_dimensions['J'].width = 15
    ws1.column_dimensions['K'].width = 30

    # ---------------------------------------------------------------------------
    # Sheet 2: Report Summary
    # ---------------------------------------------------------------------------
    ws2 = wb.create_sheet("Report Summary")

    headers2 = [
        "Report Name", "Report Description", "Primary Purpose",
        "# Columns", "Source Tables", "Business Domains"
    ]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for result in results:
        report_name = result['report_name']
        l6_data = result.get('l6_data', {})
        summary = l6_data.get('summary', {}) if l6_data else {}

        ws2.cell(row=row, column=1, value=report_name)
        ws2.cell(row=row, column=2, value=summary.get('query_summary', ''))
        ws2.cell(row=row, column=3, value=summary.get('primary_purpose', ''))
        ws2.cell(row=row, column=4, value=summary.get('column_count', len(result['l5_data'].get('columns', []))))
        ws2.cell(row=row, column=5, value=", ".join(summary.get('source_tables', [])))
        ws2.cell(row=row, column=6, value=", ".join(summary.get('business_domains', [])))
        row += 1

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 60
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 40
    ws2.column_dimensions['F'].width = 30

    # ---------------------------------------------------------------------------
    # Sheet 3: Statistics
    # ---------------------------------------------------------------------------
    ws3 = wb.create_sheet("Statistics")

    stats = [
        ("Total Reports", len(results)),
        ("Total Business Logic Terms", len(groups)),
        ("", ""),
        ("CONFLICTS (same name, different logic)", sum(1 for g in groups if g.status == "CONFLICT")),
        ("SIMILAR (needs review)", sum(1 for g in groups if g.status == "SIMILAR")),
        ("CONSISTENT (duplicates OK)", sum(1 for g in groups if g.status == "CONSISTENT")),
        ("UNIQUE (single definition)", sum(1 for g in groups if g.status == "UNIQUE")),
        ("", ""),
        ("Terms requiring steward review", sum(1 for g in groups if g.status in ("CONFLICT", "SIMILAR"))),
    ]

    for row, (label, value) in enumerate(stats, 1):
        ws3.cell(row=row, column=1, value=label)
        ws3.cell(row=row, column=2, value=value)
        if label and "CONFLICT" in label:
            ws3.cell(row=row, column=1).fill = conflict_fill
        elif label and "SIMILAR" in label:
            ws3.cell(row=row, column=1).fill = similar_fill
        elif label and "CONSISTENT" in label:
            ws3.cell(row=row, column=1).fill = consistent_fill

    ws3.column_dimensions['A'].width = 45
    ws3.column_dimensions['B'].width = 15

    # Save
    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")


def export_to_csv(results: list[dict], groups: list[BusinessLogicGroup],
                  comparison_report, output_path: str):
    """Fallback: Export to CSV if openpyxl not available."""
    import csv

    csv_path = output_path.replace('.xlsx', '.csv')

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Header
        writer.writerow([
            "Business Logic Term", "Status", "# Variations", "Report Name",
            "Business Definition", "Technical Definition", "Source Tables",
            "Business Domain", "Assigned To", "Review Status", "Notes"
        ])

        # Data
        for group in groups:
            for i, defn in enumerate(group.definitions):
                writer.writerow([
                    group.term_name if i == 0 else "",
                    group.status if i == 0 else "",
                    group.variation_count if i == 0 else "",
                    defn['report_name'],
                    defn.get('business_definition', ''),
                    defn.get('technical_definition', '')[:500],
                    ", ".join(defn.get('base_tables', [])),
                    defn.get('business_domain', ''),
                    "",  # Assigned To
                    "",  # Review Status
                    "",  # Notes
                ])

    print(f"\nCSV saved to: {csv_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Batch extract business logic from SQL files for governance"
    )
    parser.add_argument("sql_folder", help="Folder containing SQL files")
    parser.add_argument("--output", "-o", default="governance_summary.xlsx",
                        help="Output Excel file path (default: governance_summary.xlsx)")
    parser.add_argument("--schema", "-s", help="Path to clarity_schema.yaml for L6 translation")
    parser.add_argument("--api-key", help="OpenAI API key for L6 translation (or set OPENAI_API_KEY)")
    parser.add_argument("--dialect", "-d", help="SQL dialect for parsing")
    parser.add_argument("--details-dir", help="Directory to save individual L5/L6 JSON files")

    args = parser.parse_args()

    # Get API key from env if not provided
    api_key = args.api_key or os.environ.get('OPENAI_API_KEY')

    print("=" * 70)
    print("SQL BUSINESS LOGIC EXTRACTOR - GOVERNANCE BATCH PROCESS")
    print("=" * 70)
    print(f"\nInput folder: {args.sql_folder}")
    print(f"Output file: {args.output}")
    if args.schema:
        print(f"Schema: {args.schema}")
    if api_key:
        print("LLM translation: ENABLED")
    else:
        print("LLM translation: DISABLED (no API key)")
    print()

    # Step 1: Batch process all SQL files
    print("-" * 70)
    print("STEP 1: Processing SQL files")
    print("-" * 70)

    results = batch_process(
        sql_folder=args.sql_folder,
        schema_path=args.schema,
        api_key=api_key,
        dialect=args.dialect,
        output_dir=args.details_dir,
    )

    if not results:
        print("No results to process")
        return

    # Step 2: Compare all definitions
    print("\n" + "-" * 70)
    print("STEP 2: Comparing definitions across reports")
    print("-" * 70)

    comparison_report = compare_all_definitions(results)

    print(f"\nComparison results:")
    print(f"  - Conflicts: {len(comparison_report.conflicts)}")
    print(f"  - Exact duplicates: {len(comparison_report.exact_duplicates)}")
    print(f"  - Structural matches: {len(comparison_report.structural_matches)}")
    print(f"  - Semantic matches: {len(comparison_report.semantic_matches)}")

    # Step 3: Group by business logic term
    print("\n" + "-" * 70)
    print("STEP 3: Grouping by business logic term")
    print("-" * 70)

    groups = group_by_business_logic(results, comparison_report)

    conflict_count = sum(1 for g in groups if g.status == "CONFLICT")
    similar_count = sum(1 for g in groups if g.status == "SIMILAR")
    consistent_count = sum(1 for g in groups if g.status == "CONSISTENT")
    unique_count = sum(1 for g in groups if g.status == "UNIQUE")

    print(f"\nGrouped {len(groups)} business logic terms:")
    print(f"  - CONFLICT (needs resolution): {conflict_count}")
    print(f"  - SIMILAR (needs review): {similar_count}")
    print(f"  - CONSISTENT (OK): {consistent_count}")
    print(f"  - UNIQUE (single definition): {unique_count}")

    # Step 4: Export to Excel
    print("\n" + "-" * 70)
    print("STEP 4: Exporting to spreadsheet")
    print("-" * 70)

    export_to_excel(results, groups, comparison_report, args.output)

    # Summary
    print("\n" + "=" * 70)
    print("COMPLETE")
    print("=" * 70)
    print(f"\nProcessed {len(results)} reports")
    print(f"Found {len(groups)} unique business logic terms")
    print(f"  - {conflict_count + similar_count} terms need steward review")
    print(f"\nOutput: {args.output}")
    if args.details_dir:
        print(f"Details: {args.details_dir}/details/")


if __name__ == "__main__":
    main()
